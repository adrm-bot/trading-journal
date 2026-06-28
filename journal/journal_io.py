"""journal_io.py — 활성 백엔드(Excel/Notion)에서 일지 읽기 + 의도 저장."""
import os
from datetime import datetime

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*a, **k): return False

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
XLSX_PATH = os.getenv("XLSX_PATH", os.path.join(os.path.dirname(__file__), "매매일지.xlsx"))
BACKEND = os.getenv("JOURNAL_BACKEND", "excel").strip().lower()

_TEXT = ("계획/의도", "셋업", "메모")
_NUM = ("무효선(SL의도)",)
_SEL = ("감정", "상태")


def _parse_dt(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        try:
            return datetime.fromisoformat(v.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            return None
    return None


# --- 읽기 ---------------------------------------------------------------------
def _load_excel():
    from openpyxl import load_workbook
    if not os.path.exists(XLSX_PATH):
        return []
    ws = load_workbook(XLSX_PATH).active
    header = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, r))
        d["청산시각"] = _parse_dt(d.get("청산시각"))
        out.append(d)
    return out


def _notion_headers():
    return {"Authorization": f"Bearer {os.getenv('NOTION_TOKEN','').strip()}",
            "Notion-Version": "2022-06-28", "Content-Type": "application/json"}


def _load_notion():
    import requests
    db = os.getenv("NOTION_DB_ID", "").strip()
    if not (os.getenv("NOTION_TOKEN", "").strip() and db):
        return []
    rows, payload = [], {"page_size": 100}
    while True:
        r = requests.post(f"https://api.notion.com/v1/databases/{db}/query", headers=_notion_headers(), json=payload, timeout=30)
        r.raise_for_status()
        d = r.json()
        for pg in d.get("results", []):
            row = {}
            for name, prop in pg.get("properties", {}).items():
                t = prop.get("type")
                if t == "number":
                    row[name] = prop.get("number")
                elif t == "select":
                    row[name] = (prop.get("select") or {}).get("name")
                elif t == "date":
                    row[name] = _parse_dt((prop.get("date") or {}).get("start"))
                elif t == "title":
                    row[name] = "".join(x.get("plain_text", "") for x in prop.get("title", []))
                elif t == "rich_text":
                    row[name] = "".join(x.get("plain_text", "") for x in prop.get("rich_text", []))
            rows.append(row)
        if not d.get("has_more"):
            break
        payload["start_cursor"] = d["next_cursor"]
    return rows


def load_journal():
    return _load_notion() if BACKEND == "notion" else _load_excel()


# --- 의도 저장 ----------------------------------------------------------------
def _clean(updates):
    return {k: v for k, v in updates.items() if v not in (None, "")}


def _save_excel(trade_id, updates):
    from openpyxl import load_workbook
    if not os.path.exists(XLSX_PATH):
        return False
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    header = [c.value for c in ws[1]]
    if "거래ID" not in header:
        return False
    idc = header.index("거래ID")
    for row in ws.iter_rows(min_row=2):
        if str(row[idc].value) == str(trade_id):
            for k, v in _clean(updates).items():
                if k in header:
                    row[header.index(k)].value = v
            wb.save(XLSX_PATH)
            return True
    return False


def _save_notion(trade_id, updates):
    import requests
    db = os.getenv("NOTION_DB_ID", "").strip()
    h = _notion_headers()
    q = requests.post(f"https://api.notion.com/v1/databases/{db}/query", headers=h,
                      json={"filter": {"property": "거래ID", "rich_text": {"equals": str(trade_id)}}}, timeout=30)
    q.raise_for_status()
    res = q.json().get("results", [])
    if not res:
        return False
    pid = res[0]["id"]
    props = {}
    for k, v in _clean(updates).items():
        if k in _TEXT:
            props[k] = {"rich_text": [{"text": {"content": str(v)}}]}
        elif k in _NUM:
            props[k] = {"number": float(v)}
        elif k in _SEL:
            props[k] = {"select": {"name": str(v)}}
    r = requests.patch(f"https://api.notion.com/v1/pages/{pid}", headers=h, json={"properties": props}, timeout=30)
    r.raise_for_status()
    return True


def save_intent(trade_id, updates):
    """updates: {계획/의도, 셋업, 무효선(SL의도), 감정, 메모, 상태}."""
    return _save_notion(trade_id, updates) if BACKEND == "notion" else _save_excel(trade_id, updates)
