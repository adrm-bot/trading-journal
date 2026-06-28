#!/usr/bin/env python3
"""
pull_trades.py — 거래소 청산 거래를 매매일지에 자동 적재 (read-only, 멀티거래소).

ccxt 기반 거래소-무관 엔진. EXCHANGES 에 적힌 거래소들을 순회하며 각자의 read-only
키로 청산거래를 긁어 하나의 통일 스키마로 정규화 → Excel(기본) 또는 Notion 에 적재.

거래소 추가 = 어댑터 함수 1개 (ADAPTERS 에 등록). 현재: bybit(완전), binance(v0).

설정(journal/.env):
  EXCHANGES=bybit            # 콤마로 여러개: bybit,binance
  BYBIT_API_KEY/SECRET       # 거래소별 키 (BINANCE_API_KEY/SECRET ...)
  JOURNAL_BACKEND=excel|notion
사용:
  python pull_trades.py            # 적재 (멱등)
  python pull_trades.py --pending  # 의도 미기입 목록 (excel 백엔드)
"""
from __future__ import annotations
import os, sys, time
from datetime import datetime, timezone

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*a, **k): return False
try:
    import ccxt
except ImportError:
    ccxt = None

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

EXCHANGES = [e.strip().lower() for e in os.getenv("EXCHANGES", "bybit").split(",") if e.strip()]
LOOKBACK_DAYS = int(os.getenv("LOOKBACK_DAYS", "7"))
JOURNAL_BACKEND = os.getenv("JOURNAL_BACKEND", "excel").strip().lower()
XLSX_PATH = os.getenv("XLSX_PATH", os.path.join(os.path.dirname(__file__), "매매일지.xlsx"))
STATUS_PENDING = "의도 미기입"

COLUMNS = [
    "청산시각", "거래소", "심볼", "방향", "진입가", "청산가", "수량", "실현손익(USDT)",
    "R", "수수료(USDT)", "보유시간(분)", "상태",
    "계획/의도", "셋업", "무효선(SL의도)", "감정", "메모", "출처", "거래ID",
]


def _keys(exchange: str) -> tuple[str, str]:
    k = os.getenv(f"{exchange.upper()}_API_KEY", "").strip()
    s = os.getenv(f"{exchange.upper()}_API_SECRET", "").strip()
    if not (k and s):
        sys.exit(f"{exchange.upper()}_API_KEY / _API_SECRET 미설정 (journal/.env)")
    return k, s


def _fnum(d, key):
    try:
        return float(d.get(key))
    except (TypeError, ValueError):
        return None


def _row(exchange, trade_id, updated_ms, symbol, direction, entry, exit_, qty, pnl):
    dt = datetime.fromtimestamp(updated_ms / 1000, tz=timezone.utc)
    return {
        "exchange": exchange,
        "trade_id": trade_id,
        "closed_at_dt": dt.replace(tzinfo=None),
        "closed_at_iso": dt.isoformat(),
        "symbol": symbol, "direction": direction,
        "entry": entry, "exit": exit_, "qty": qty, "pnl": pnl,
    }


# --- 거래소 어댑터 ------------------------------------------------------------
def fetch_bybit(lookback_days: int) -> list[dict]:
    """Bybit V5 /v5/position/closed-pnl (ccxt implicit). 포지션별 평균진입/청산."""
    key, secret = _keys("bybit")
    ex = ccxt.bybit({"apiKey": key, "secret": secret, "enableRateLimit": True})
    now = ex.milliseconds()
    start = now - lookback_days * 86_400_000
    chunk = 7 * 86_400_000
    rows, win = [], start
    while win < now:
        end = min(win + chunk, now)
        cursor = None
        while True:
            p = {"category": "linear", "startTime": win, "endTime": end, "limit": 100}
            if cursor:
                p["cursor"] = cursor
            resp = ex.private_get_v5_position_closed_pnl(p)
            if str(resp.get("retCode")) != "0":
                sys.exit(f"Bybit 오류: {resp.get('retCode')} {resp.get('retMsg')}")
            res = resp.get("result", {})
            for r in res.get("list", []):
                sym = r.get("symbol", "")
                oid = r.get("orderId", "")
                upd = int(r.get("updatedTime") or r.get("createdTime") or 0)
                rows.append(_row(
                    "bybit", f"bybit:{sym}:{oid}:{upd}", upd, sym,
                    "Long" if r.get("side") == "Sell" else "Short",
                    _fnum(r, "avgEntryPrice"), _fnum(r, "avgExitPrice"),
                    _fnum(r, "qty") or _fnum(r, "closedSize"), _fnum(r, "closedPnl"),
                ))
            cursor = res.get("nextPageCursor")
            if not cursor:
                break
            time.sleep(0.2)
        win = end
    return rows


def fetch_binance(lookback_days: int) -> list[dict]:
    """Binance USDⓈ-M income(REALIZED_PNL). v0: 실현손익·심볼·시각만(진입/청산가는 후속)."""
    key, secret = _keys("binance")
    ex = ccxt.binanceusdm({"apiKey": key, "secret": secret, "enableRateLimit": True})
    now = ex.milliseconds()
    start = now - lookback_days * 86_400_000
    rows, cur = [], start
    week = 7 * 86_400_000
    while cur < now:
        end = min(cur + week, now)
        incs = ex.fapiPrivateGetIncome({
            "incomeType": "REALIZED_PNL", "startTime": cur, "endTime": end, "limit": 1000,
        })
        for inc in incs:
            pnl = _fnum(inc, "income")
            if not pnl:
                continue
            sym = inc.get("symbol", "")
            t = int(inc.get("time", 0))
            tran = inc.get("tranId", "")
            rows.append(_row(
                "binance", f"binance:{sym}:{tran}:{t}", t, sym,
                None, None, None, None, pnl,  # 방향/진입/청산은 v0 미수집
            ))
        cur = end
    return rows


ADAPTERS = {"bybit": fetch_bybit, "binance": fetch_binance}


# --- Excel 백엔드 -------------------------------------------------------------
def _excel_load():
    from openpyxl import Workbook, load_workbook
    if os.path.exists(XLSX_PATH):
        wb = load_workbook(XLSX_PATH); ws = wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = "매매일지"
        ws.append(COLUMNS); ws.freeze_panes = "A2"
    return wb, ws


def _excel_ids(ws):
    idx = COLUMNS.index("거래ID") + 1
    return {str(r[0]) for r in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True) if r[0]}


def excel_upsert(rows):
    wb, ws = _excel_load()
    existing = _excel_ids(ws)
    added = skipped = 0
    for r in rows:
        if r["trade_id"] in existing:
            skipped += 1; continue
        ws.append([
            r["closed_at_dt"], r["exchange"], r["symbol"], r["direction"], r["entry"], r["exit"],
            r["qty"], r["pnl"], None, None, None, STATUS_PENDING,
            None, None, None, None, None, "manual", r["trade_id"],
        ])
        ws.cell(row=ws.max_row, column=1).number_format = "yyyy-mm-dd hh:mm"
        existing.add(r["trade_id"]); added += 1
        print(f"  + [{r['exchange']}] {r['symbol']} {r['direction'] or ''} {r['closed_at_iso'][:16]}  (손익 {r['pnl']})")
    try:
        wb.save(XLSX_PATH)
    except PermissionError:
        sys.exit(f"저장 실패: 엑셀에서 '{os.path.basename(XLSX_PATH)}' 닫고 재실행.")
    return added, skipped


def excel_pending():
    if not os.path.exists(XLSX_PATH):
        print("일지 파일 없음."); return
    from openpyxl import load_workbook
    ws = load_workbook(XLSX_PATH).active
    s = COLUMNS.index("상태")
    pend = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[s] == STATUS_PENDING]
    if not pend:
        print("의도 미기입 없음."); return
    print(f"의도 미기입 {len(pend)}건:")
    for r in pend:
        print(f"  - [{r[COLUMNS.index('거래소')]}] {r[COLUMNS.index('심볼')]} {r[COLUMNS.index('방향')] or ''} {r[0]} 손익={r[COLUMNS.index('실현손익(USDT)')]}")


# --- Notion 백엔드 ------------------------------------------------------------
def notion_upsert(rows):
    import requests
    token = os.getenv("NOTION_TOKEN", "").strip()
    db_id = os.getenv("NOTION_DB_ID", "").strip()
    if not (token and db_id):
        sys.exit("NOTION_TOKEN / NOTION_DB_ID 미설정 (journal/.env)")
    h = {"Authorization": f"Bearer {token}", "Notion-Version": "2022-06-28", "Content-Type": "application/json"}
    existing, payload = set(), {"page_size": 100}
    while True:
        r = requests.post(f"https://api.notion.com/v1/databases/{db_id}/query", headers=h, json=payload, timeout=30)
        r.raise_for_status(); d = r.json()
        for pg in d.get("results", []):
            rich = pg.get("properties", {}).get("거래ID", {}).get("rich_text", [])
            if rich:
                existing.add(rich[0].get("plain_text", ""))
        if not d.get("has_more"):
            break
        payload["start_cursor"] = d["next_cursor"]
    added = skipped = 0
    for r in rows:
        if r["trade_id"] in existing:
            skipped += 1; continue
        props = {
            "거래": {"title": [{"text": {"content": f"[{r['exchange']}] {r['symbol']} {r['direction'] or ''} {r['closed_at_iso'][:16]}"}}]},
            "거래ID": {"rich_text": [{"text": {"content": r["trade_id"]}}]},
            "거래소": {"select": {"name": r["exchange"]}},
            "심볼": {"rich_text": [{"text": {"content": r["symbol"]}}]},
            "청산시각": {"date": {"start": r["closed_at_iso"]}},
            "상태": {"select": {"name": STATUS_PENDING}},
            "출처": {"select": {"name": "manual"}},
        }
        if r["direction"]:
            props["방향"] = {"select": {"name": r["direction"]}}
        for k, key in (("진입가", "entry"), ("청산가", "exit"), ("수량", "qty"), ("실현손익(USDT)", "pnl")):
            if r.get(key) is not None:
                props[k] = {"number": r[key]}
        resp = requests.post("https://api.notion.com/v1/pages", headers=h,
                             json={"parent": {"database_id": db_id}, "properties": props}, timeout=30)
        resp.raise_for_status(); added += 1
        print(f"  + [{r['exchange']}] {r['symbol']} {r['closed_at_iso'][:16]}")
    return added, skipped


# --- main ---------------------------------------------------------------------
def main():
    if ccxt is None:
        sys.exit("ccxt 미설치: pip install -r journal/requirements.txt")
    if "--pending" in sys.argv:
        excel_pending() if JOURNAL_BACKEND == "excel" else print("--pending 은 excel 전용.")
        return

    rows = []
    for ex in EXCHANGES:
        if ex not in ADAPTERS:
            print(f"  ! 미지원 거래소 '{ex}' — 건너뜀 (지원: {', '.join(ADAPTERS)})"); continue
        n = len(rows)
        rows += ADAPTERS[ex](LOOKBACK_DAYS)
        print(f"  · {ex}: {len(rows)-n}건 조회")
    rows.sort(key=lambda r: r["closed_at_iso"], reverse=True)

    if JOURNAL_BACKEND == "notion":
        added, skipped = notion_upsert(rows); where = "Notion"
    else:
        added, skipped = excel_upsert(rows); where = XLSX_PATH
    print(f"\n완료: 신규 {added} / skip {skipped} / 총 {len(rows)} → {where}")
    if added:
        print("→ '의도 미기입' 거래의 계획을 사후에 채워라.")


if __name__ == "__main__":
    main()
